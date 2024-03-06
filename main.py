import openpyxl

# Excel dosyası adı
excel_dosyasi = "datas/donusumler.xlsx"

# Giriş dosyası adı
giris_dosyasi = "datas/inputs/giris.txt"

# Çıkış dosyası adı
cikis_dosyasi = "datas/outputs/cikis.txt"

try:
    # Excel dosyasını açıyoruz
    workbook = openpyxl.load_workbook(excel_dosyasi)
    sheet = workbook.active

    # Kolon adları ve dönüşümleri için bir sözlük oluşturuyoruz
    donusumler = {}
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=2):
        kolon_adi = row[0].value
        donusum = row[1].value
        donusumler[kolon_adi] = donusum

    # Giriş ve çıkış dosyalarını açıyoruz
    with open(giris_dosyasi, "r") as giris, open(cikis_dosyasi, "w") as cikis:
        # Giriş dosyasını satır satır okuyoruz
        for satir in giris:
            # Her satırı boşluklardan temizliyoruz
            satir = satir.strip()
            # Eğer satır boş ise atlıyoruz
            if not satir:
                continue
            # Kolon adını alıyoruz
            kolon_adi = satir.split(",")[0].strip()
            # Eğer kolon adı donüşümler sözlüğünde varsa, dönüşümü uyguluyoruz
            if kolon_adi in donusumler:
                dönüşüm = donusumler[kolon_adi]
                cikis.write(dönüşüm + ",\n")

    print("İşlem tamamlandı. Çıkış dosyası '{}' adında oluşturuldu.".format(cikis_dosyasi))

except FileNotFoundError:
    print("Dosya bulunamadı.")

except Exception as e:
    print("Bir hata oluştu:", e)